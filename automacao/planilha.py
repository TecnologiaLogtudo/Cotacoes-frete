from __future__ import annotations

import gc
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List
import unicodedata
import re

from openpyxl import load_workbook


@dataclass
class LinhaAutomacao:
    numero: str
    nome_motorista: str
    placa: str
    perfil: str
    base: str
    operacao: str
    frete_negociado: str
    excel_row: int


def _normalizar(texto: str) -> str:
    sem_acentos = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII")
    return sem_acentos.strip().upper()


def _texto_celula(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%d/%m/%Y")
    return str(valor).strip()


def _formatar_moeda(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, (int, float)):
        return f"{valor:.2f}".replace(".", ",")

    texto = str(valor).strip()
    if not texto:
        return ""

    return texto


def _mapear_headers(ws, header_row: int) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        nome = _texto_celula(ws.cell(row=header_row, column=col).value)
        if nome:
            headers[_normalizar(nome)] = col
    return headers


def _encontrar_coluna_data(ws, header_row: int, data_referencia: str) -> int:
    data_norm = _normalizar(data_referencia)
    data_ref_parts = _extrair_partes_data(data_referencia)
    for col in range(1, ws.max_column + 1):
        valor_cabecalho = ws.cell(row=header_row, column=col).value
        cabecalho = _texto_celula(valor_cabecalho)
        if cabecalho and _normalizar(cabecalho) == data_norm:
            return col

        # Match flexível para cenários como:
        # - cabeçalho exibido como "27-mar"
        # - valor interno/formula exibido como "27/03/2026"
        if data_ref_parts:
            cabecalho_parts = _extrair_partes_data(valor_cabecalho)
            if _mesma_data(data_ref_parts, cabecalho_parts):
                return col

    cabecalhos_encontrados: List[str] = []
    for col in range(1, ws.max_column + 1):
        texto = _texto_celula(ws.cell(row=header_row, column=col).value)
        if texto:
            cabecalhos_encontrados.append(texto)
    resumo_cabecalhos = ", ".join(cabecalhos_encontrados[:12]) if cabecalhos_encontrados else "(vazio)"

    raise ValueError(
        f"Nao foi encontrada coluna de data '{data_referencia}' na linha de cabecalho {header_row}. "
        f"Cabecalhos encontrados: {resumo_cabecalhos}"
    )


def _extrair_partes_data(valor: Any) -> tuple[int, int, int | None] | None:
    if valor is None:
        return None

    if isinstance(valor, datetime):
        return (valor.day, valor.month, valor.year)
    if isinstance(valor, date):
        return (valor.day, valor.month, valor.year)

    texto = str(valor).strip()
    if not texto:
        return None

    # dd/mm/yyyy, dd-mm-yyyy, dd.mm.yyyy
    m = re.fullmatch(r"(\d{1,2})[\/\-.](\d{1,2})[\/\-.](\d{2,4})", texto)
    if m:
        dia = int(m.group(1))
        mes = int(m.group(2))
        ano = int(m.group(3))
        if ano < 100:
            ano += 2000
        return (dia, mes, ano)

    # yyyy-mm-dd
    m = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", texto)
    if m:
        ano = int(m.group(1))
        mes = int(m.group(2))
        dia = int(m.group(3))
        return (dia, mes, ano)

    # dd-mmm (pt-BR/pt-PT), ex: 27-mar
    m = re.fullmatch(r"(\d{1,2})\s*[-\/]\s*([a-zA-ZçÇ]{3,9})", texto, flags=re.IGNORECASE)
    if m:
        dia = int(m.group(1))
        mes_txt = _normalizar(m.group(2))
        meses = {
            "JAN": 1,
            "FEV": 2,
            "FEB": 2,
            "MAR": 3,
            "ABR": 4,
            "APR": 4,
            "MAI": 5,
            "MAY": 5,
            "JUN": 6,
            "JUL": 7,
            "AGO": 8,
            "AUG": 8,
            "SET": 9,
            "SEP": 9,
            "OUT": 10,
            "OCT": 10,
            "NOV": 11,
            "DEZ": 12,
            "DEC": 12,
        }
        mes = meses.get(mes_txt[:3])
        if mes:
            return (dia, mes, None)

    return None


def _mesma_data(
    referencia: tuple[int, int, int | None] | None,
    candidata: tuple[int, int, int | None] | None,
) -> bool:
    if not referencia or not candidata:
        return False

    dia_ref, mes_ref, ano_ref = referencia
    dia_can, mes_can, ano_can = candidata

    if dia_ref != dia_can or mes_ref != mes_can:
        return False

    # Quando o cabeçalho vem como "dd-mmm", não há ano explícito.
    if ano_ref is None or ano_can is None:
        return True

    return ano_ref == ano_can


def _encontrar_coluna_numero(ws, header_row: int) -> int:
    headers = _mapear_headers(ws, header_row=header_row)
    col_numero = headers.get(_normalizar("Número")) or headers.get(_normalizar("Numero"))
    if not col_numero:
        raise ValueError("Cabecalho 'Número' nao encontrado na linha 3.")
    return col_numero


def carregar_linhas_para_automacao(
    planilha_path: str,
    data_referencia: str,
    header_row: int = 3,
    max_rows_to_scan: int = 100,
) -> List[LinhaAutomacao]:
    arquivo = Path(planilha_path)
    if not arquivo.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {planilha_path}")

    wb = load_workbook(filename=str(arquivo), data_only=True, read_only=True)
    ws = wb.active
    linhas: List[LinhaAutomacao] = []

    try:
        headers = _mapear_headers(ws, header_row=header_row)
        col_numero = headers.get(_normalizar("Número")) or headers.get(_normalizar("Numero"))
        if not col_numero:
            raise ValueError("Cabecalho 'Número' nao encontrado na linha 3.")

        col_nome = headers.get(_normalizar("Nome"), 1)
        col_placa = headers.get(_normalizar("Placa"), 2)
        col_perfil = headers.get(_normalizar("Perfil"))
        col_base = headers.get(_normalizar("Base"))
        col_operacao = headers.get(_normalizar("Operação")) or headers.get(_normalizar("Operacao"))

        if not col_perfil:
            raise ValueError("Cabecalho 'Perfil' nao encontrado na linha 3.")
        if not col_base:
            raise ValueError("Cabecalho 'Base' nao encontrado na linha 3.")

        col_data = _encontrar_coluna_data(ws, header_row=header_row, data_referencia=data_referencia)

        inicio = header_row + 1
        fim = inicio + max_rows_to_scan - 1

        for idx, values in enumerate(
            ws.iter_rows(min_row=inicio, max_row=fim, values_only=True),
            start=inicio,
        ):
            numero = _texto_celula(values[col_numero - 1] if len(values) >= col_numero else "")
            if not numero:
                continue

            nome = _texto_celula(values[col_nome - 1] if len(values) >= col_nome else "")
            placa = _texto_celula(values[col_placa - 1] if len(values) >= col_placa else "")
            perfil = _texto_celula(values[col_perfil - 1] if len(values) >= col_perfil else "")
            base = _texto_celula(values[col_base - 1] if len(values) >= col_base else "")
            operacao = _texto_celula(values[col_operacao - 1] if col_operacao and len(values) >= col_operacao else "")
            frete = _formatar_moeda(values[col_data - 1] if len(values) >= col_data else "")

            linhas.append(
                LinhaAutomacao(
                    numero=numero,
                    nome_motorista=nome,
                    placa=placa,
                    perfil=perfil,
                    base=base,
                    operacao=operacao,
                    frete_negociado=frete,
                    excel_row=idx,
                )
            )
    finally:
        wb.close()
        gc.collect()

    return linhas
